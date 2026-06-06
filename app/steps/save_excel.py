from __future__ import annotations

import os
import re
import shutil
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from app.pipeline.context import RunContext
from app.pipeline.registry import REGISTRY, StepDefinition
from app.pipeline.schema import Step
from app.steps.util import ensure_df_exists, get_required_param, resolve_df_names_by_mask


def _safe_filename(name: str) -> str:
    name = str(name)
    name = re.sub(r"[<>:\"/\\\\|?*]+", "_", name)
    name = name.strip().strip(".")
    return name or "output"


def _normalize_export_mode(raw: Any) -> str:
    s = str(raw or "single").strip().lower()
    if s in ("single", "one", "source"):
        return "single"
    if s in ("mask_sheets", "by_mask", "mask", "multi_df", "multi"):
        return "mask_sheets"
    raise ValueError(
        "save_excel: export_mode must be single (один source_df) or "
        "mask_sheets (несколько DF по маске, каждый на свой лист)"
    )


def _parse_columns_param(raw: Any) -> list[str]:
    columns = raw or []
    if isinstance(columns, str):
        return [c.strip() for c in columns.split(",") if c.strip()]
    if isinstance(columns, list):
        return [str(c).strip() for c in columns if str(c).strip()]
    return []


def _subset_df_columns(df: pd.DataFrame, columns: list[str], *, df_label: str) -> pd.DataFrame:
    if not columns:
        return df
    miss = [c for c in columns if c not in df.columns]
    if miss:
        raise ValueError(f"save_excel: columns not found in {df_label}: {miss}")
    return df.loc[:, list(columns)]


def _safe_excel_sheet_name(name: str, used: set[str]) -> str:
    """Имя листа Excel: до 31 символа, без []:*?/\\."""
    s = re.sub(r"[\[\]:*?/\\]+", "_", str(name).strip())
    s = s[:31] if s else "Sheet"
    base = s
    n = 1
    while s in used:
        suffix = f"_{n}"
        max_base = 31 - len(suffix)
        s = (base[:max_base] if max_base > 0 else "") + suffix
        if not s:
            s = f"Sheet{n}"[:31]
        n += 1
    used.add(s)
    return s


def _write_df_into_template(
    template_path: str,
    sheet_name: str,
    df: pd.DataFrame,
    out_path: str,
    start_row: int,
    start_col: int,
    column_map: dict[str, int] | None = None,
    *,
    write_mode: str = "overwrite",
) -> None:
    mode = str(write_mode or "overwrite").strip().lower()
    if mode not in ("overwrite", "update"):
        raise ValueError("template_write_mode must be overwrite or update")

    # overwrite: create out_path as a copy of template
    # update: open existing out_path and write into it (no template copy)
    if mode == "overwrite":
        if not os.path.isfile(template_path):
            raise FileNotFoundError(f"save_excel: шаблон не найден: {template_path}")
        os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
        shutil.copy2(template_path, out_path)
        wb = load_workbook(out_path)
    else:
        if not os.path.isfile(out_path):
            raise FileNotFoundError(
                f"Template update mode requires existing file: {out_path}"
            )
        wb = load_workbook(out_path)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif sheet_name:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb.active

    r0 = max(1, start_row)
    c0 = max(1, start_col)

    if column_map:
        # Write each dataframe column into its mapped absolute Excel column.
        for r_idx in range(len(df)):
            for col_name in df.columns:
                if col_name not in column_map:
                    continue
                excel_col = int(column_map[col_name])
                val = df.iloc[r_idx][col_name]
                ws.cell(
                    row=r0 + r_idx,
                    column=excel_col,
                    value=None if pd.isna(val) else val,
                )
    else:
        for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=0):
            for c_idx, val in enumerate(row, start=0):
                ws.cell(
                    row=r0 + r_idx,
                    column=c0 + c_idx,
                    value=None if pd.isna(val) else val,
                )

    wb.save(out_path)


def _write_plain_dataframe_excel(
    out_path: str,
    df: pd.DataFrame,
    sheet_name: str,
    start_row: int,
    start_col: int,
    *,
    writer_mode: str,
    if_sheet_exists: str,
) -> None:
    """
    Запись DF без шаблона через pandas.ExcelWriter + openpyxl.

    writer_mode:
      - w — новый файл / полная перезапись (как раньше по умолчанию).
      - a — дописать в существующий .xlsx; если файла ещё нет, создаётся через mode=w.

    if_sheet_exists — только для mode=a и существующего файла (pandas/openpyxl):
      replace | overlay | error | new
    """
    wm = str(writer_mode or "w").strip().lower()
    if wm not in ("w", "a"):
        raise ValueError("save_excel: writer_mode must be w or a")

    ise = str(if_sheet_exists or "replace").strip().lower()
    if ise not in ("replace", "overlay", "error", "new"):
        raise ValueError(
            "save_excel: if_sheet_exists must be one of: replace, overlay, error, new"
        )

    startrow0 = max(0, int(start_row) - 1)
    startcol0 = max(0, int(start_col) - 1)

    def _to_excel(writer: pd.ExcelWriter) -> None:
        df.to_excel(
            writer,
            index=False,
            sheet_name=sheet_name,
            startrow=startrow0,
            startcol=startcol0,
        )

    if wm == "w":
        with pd.ExcelWriter(out_path, engine="openpyxl", mode="w") as writer:
            _to_excel(writer)
        return

    # append: файл должен существовать, иначе pandas mode='a' падает — создаём как w
    if not os.path.isfile(out_path):
        with pd.ExcelWriter(out_path, engine="openpyxl", mode="w") as writer:
            _to_excel(writer)
        return

    with pd.ExcelWriter(
        out_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists=ise,
    ) as writer:
        _to_excel(writer)


def _write_plain_multi_excel(
    out_path: str,
    sheets: list[tuple[str, pd.DataFrame]],
    start_row: int,
    start_col: int,
    *,
    writer_mode: str,
    if_sheet_exists: str,
) -> None:
    """
    Несколько DF в один .xlsx без шаблона: каждый элемент — (имя листа, DataFrame).

    Имена листов должны быть уже уникальны и допустимы для Excel.
    """
    if not sheets:
        raise ValueError("save_excel mask_sheets: нет листов для записи")

    wm = str(writer_mode or "w").strip().lower()
    if wm not in ("w", "a"):
        raise ValueError("save_excel: writer_mode must be w or a")

    ise = str(if_sheet_exists or "replace").strip().lower()
    if ise not in ("replace", "overlay", "error", "new"):
        raise ValueError(
            "save_excel: if_sheet_exists must be one of: replace, overlay, error, new"
        )

    startrow0 = max(0, int(start_row) - 1)
    startcol0 = max(0, int(start_col) - 1)

    def _write_all(writer: pd.ExcelWriter) -> None:
        for sheet_name, frame in sheets:
            frame.to_excel(
                writer,
                index=False,
                sheet_name=sheet_name,
                startrow=startrow0,
                startcol=startcol0,
            )

    if wm == "w" or not os.path.isfile(out_path):
        with pd.ExcelWriter(out_path, engine="openpyxl", mode="w") as writer:
            _write_all(writer)
        return

    with pd.ExcelWriter(
        out_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists=ise,
    ) as writer:
        _write_all(writer)


def run_save_excel(ctx: RunContext, step: Step) -> None:
    p = step.params
    export_mode = _normalize_export_mode(p.get("export_mode", "single"))
    out_dir = str(p.get("out_dir", "") or "").strip()
    if not out_dir:
        raise ValueError("save_excel: задайте out_dir или добавьте диалог в params.dialogs")
    filename = str(p.get("filename", "") or "").strip() or "result.xlsx"

    columns = _parse_columns_param(p.get("columns"))

    os.makedirs(out_dir, exist_ok=True)
    sheet_name = str(p.get("sheet_name", "Sheet1"))
    start_row = int(p.get("start_row", 1))
    start_col = int(p.get("start_col", 1))
    template_path = str(p.get("template_path", "") or "").strip()
    template_write_mode = str(p.get("template_write_mode", "overwrite") or "overwrite").strip().lower()

    writer_mode = str(p.get("writer_mode", "w") or "w").strip().lower()
    if_sheet_exists = str(p.get("if_sheet_exists", "replace") or "replace").strip().lower()

    split_by = str(p.get("split_by_column", "") or "").strip()
    split_mask = str(p.get("split_filename_mask", "{group}.xlsx") or "{group}.xlsx")

    if export_mode == "mask_sheets":
        if template_path:
            raise ValueError(
                "save_excel mask_sheets: template_path не поддерживается; оставьте пустым"
            )
        if split_by:
            raise ValueError(
                "save_excel mask_sheets: split_by_column несовместим с режимом mask_sheets"
            )

    template_column_map = p.get("template_column_map") or {}
    if isinstance(template_column_map, list):
        # allow list of {df_col, excel_col}
        m: dict[str, int] = {}
        for item in template_column_map:
            if isinstance(item, dict) and "df_col" in item and "excel_col" in item:
                m[str(item["df_col"])] = int(item["excel_col"])
        template_column_map = m
    if not isinstance(template_column_map, dict):
        template_column_map = {}

    def _is_file_locked_error(e: BaseException) -> bool:
        # Windows: PermissionError / OSError with winerror=32 is typical for "file in use".
        if isinstance(e, PermissionError):
            return True
        if isinstance(e, OSError):
            winerr = getattr(e, "winerror", None)
            if winerr == 32:
                return True
        return False

    def _ask_retry_locked(path: str, err: BaseException) -> bool:
        fn = ctx.variables.get("tk_askretrycancel")
        if not callable(fn):
            return False
        return bool(
            fn(
                "ExcelForge",
                "Файл занят другим процессом и недоступен для записи.\n\n"
                f"Файл:\n{path}\n\n"
                f"Ошибка:\n{err}\n\n"
                "Закройте файл (например, в Excel) и нажмите Retry, чтобы повторить.",
            )
        )

    def _save_mask_sheets(out_path: str) -> None:
        names = resolve_df_names_by_mask(ctx.df_store, p, step_label="save_excel")
        used_sheets: set[str] = set()
        sheets: list[tuple[str, pd.DataFrame]] = []
        for df_name in names:
            frame = _subset_df_columns(ctx.df_store[df_name], columns, df_label=df_name)
            sheet = _safe_excel_sheet_name(df_name, used_sheets)
            sheets.append((sheet, frame))

        while True:
            try:
                ctx.logger.info(
                    "save_excel mask_sheets: "
                    f"writer_mode={writer_mode}, if_sheet_exists={if_sheet_exists}, "
                    f"sheets={len(sheets)}, path={out_path}, "
                    f"dfs={', '.join(names)}"
                )
                _write_plain_multi_excel(
                    out_path,
                    sheets,
                    start_row,
                    start_col,
                    writer_mode=writer_mode,
                    if_sheet_exists=if_sheet_exists,
                )
                return
            except Exception as e:  # noqa: BLE001
                if _is_file_locked_error(e) and _ask_retry_locked(out_path, e):
                    continue
                raise

    def _save_one(one_df: pd.DataFrame, out_path: str) -> None:
        while True:
            try:
                if template_path:
                    _write_df_into_template(
                        template_path=template_path,
                        sheet_name=sheet_name,
                        df=one_df,
                        out_path=out_path,
                        start_row=start_row,
                        start_col=start_col,
                        column_map={str(k): int(v) for k, v in template_column_map.items()}
                        if template_column_map
                        else None,
                        write_mode=template_write_mode,
                    )
                else:
                    ctx.logger.info(
                        "save_excel plain: "
                        f"writer_mode={writer_mode}, if_sheet_exists={if_sheet_exists}, "
                        f"sheet={sheet_name!r}, path={out_path}"
                    )
                    _write_plain_dataframe_excel(
                        out_path,
                        one_df,
                        sheet_name,
                        start_row,
                        start_col,
                        writer_mode=writer_mode,
                        if_sheet_exists=if_sheet_exists,
                    )
                return
            except Exception as e:  # noqa: BLE001
                if _is_file_locked_error(e) and _ask_retry_locked(out_path, e):
                    continue
                raise

    if export_mode == "mask_sheets":
        out_path = os.path.join(out_dir, filename)
        _save_mask_sheets(out_path)
        ctx.logger.info(f"Saved Excel (mask_sheets): {out_path}")
        return

    source_df = str(get_required_param(p, "source_df"))
    df = ensure_df_exists(ctx.df_store, source_df)
    df_out = _subset_df_columns(df, columns, df_label=source_df)

    if split_by:
        if split_by not in df_out.columns:
            raise ValueError(f"split_by_column not found: {split_by}")
        for group_value, gdf in df_out.groupby(split_by, dropna=False):
            group_str = _safe_filename(group_value)
            base = split_mask.format(group=group_str)
            out_path = os.path.join(out_dir, base)
            _save_one(gdf, out_path)
            ctx.logger.info(
                f"Saved group Excel: {out_path} group={group_value} rows={len(gdf)}"
            )
    else:
        out_path = os.path.join(out_dir, filename)
        _save_one(df_out, out_path)
        ctx.logger.info(
            f"Saved Excel: {out_path} rows={len(df_out)} cols={len(df_out.columns)}"
        )


def register_save_excel() -> None:
    REGISTRY.register(
        StepDefinition(
            type="save_excel",
            title="Выгрузка DataFrame → Excel",
            runner=run_save_excel,
            default_params={
                "export_mode": "single",  # single | mask_sheets
                "source_df": "df_main",
                "dataframes": [],
                "name_glob": "",
                "out_dir": "",
                "filename": "result.xlsx",
                "dialogs": [],
                "directory_initial": "",
                "columns": [],
                "template_path": "",
                "template_write_mode": "overwrite",  # overwrite|update
                "template_column_map": {},
                "sheet_name": "Sheet1",
                "start_row": 1,
                "start_col": 1,
                # plain write (no template): w = overwrite file; a = append / merge workbook
                "writer_mode": "w",  # w|a
                "if_sheet_exists": "replace",  # replace|overlay|error|new (only with writer_mode a)
                "split_by_column": "",
                "split_filename_mask": "{group}.xlsx",
            },
        )
    )

