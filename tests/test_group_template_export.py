from __future__ import annotations

import os
import tempfile
import unittest

import pandas as pd
from openpyxl import load_workbook

from app.steps.excel_template import (
    build_output_filename,
    replace_placeholders_in_text,
    safe_eval_scalar_expr,
    write_form_row_to_template,
    write_group_to_template,
)
from app.steps.group_template_export import (
    _compute_aggregations,
    _is_single_row_form_mode,
    _parse_group_by,
    _resolve_filename_params,
)


class GroupTemplateExportTests(unittest.TestCase):
    def test_parse_group_by(self) -> None:
        self.assertEqual(_parse_group_by("A"), ["A"])
        self.assertEqual(_parse_group_by(["A", "B"]), ["A", "B"])

    def test_build_filename(self) -> None:
        name = build_output_filename(prefix="Act_", group_part="001", suffix="_Q1")
        self.assertTrue(name.startswith("Act_"))
        self.assertTrue(name.endswith(".xlsx"))

    def test_filename_mask(self) -> None:
        p = {"filename_mask": "Report_{group}_final", "extension": ".xlsx"}
        name = _resolve_filename_params(p, "East")
        self.assertEqual(name, "Report_East_final.xlsx")

    def test_safe_eval_expr(self) -> None:
        self.assertEqual(safe_eval_scalar_expr("a + b * 2", {"a": 10.0, "b": 5.0}), 20.0)

    def test_aggregations(self) -> None:
        gdf = pd.DataFrame({"Qty": ["1", "2", "3"], "Price": [10, 20, 30]})
        specs = [
            {"name": "total_qty", "op": "sum", "column": "Qty"},
            {"name": "total_price", "op": "sum", "column": "Price"},
            {"name": "avg_price", "op": "expr", "expression": "total_price / total_qty"},
        ]
        res = _compute_aggregations(gdf, specs, {})
        self.assertEqual(res["total_qty"], 6.0)
        self.assertEqual(res["total_price"], 60.0)
        self.assertEqual(res["avg_price"], 10.0)

    def test_placeholders(self) -> None:
        text = replace_placeholders_in_text(
            "Group {{group.X}} sum={{agg.total}} period={{@period}}",
            group_values={"X": "ABC"},
            agg_values={"total": 100},
            var_values={"period": "2025"},
        )
        self.assertIn("ABC", text)
        self.assertIn("100", text)
        self.assertIn("2025", text)

    def test_filename_inc(self) -> None:
        name = build_output_filename(prefix="Act_", group_part="Sales", inc=3, inc_position="prefix")
        self.assertTrue(name.startswith("3_Act_"))

        name_suffix = build_output_filename(prefix="Act_", group_part="Sales", inc=3, inc_position="suffix")
        self.assertEqual(name_suffix, "Act_Sales_3.xlsx")

        p = {"prefix": "Doc_", "extension": ".xlsx", "filename_inc": False}
        name2 = _resolve_filename_params(p, "East", inc=2, inc_position=None)
        self.assertEqual(name2, "Doc_East.xlsx")

    def test_form_mode_detection(self) -> None:
        self.assertTrue(_is_single_row_form_mode({"table_start_row": "", "table_columns": []}))
        self.assertTrue(_is_single_row_form_mode({"single_row_mode": True, "table_start_row": 10}))
        self.assertFalse(
            _is_single_row_form_mode({"table_start_row": 10, "table_columns": [{"df_col": "A", "excel_col": 1}]})
        )

    def test_write_form_row_template(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tpl = os.path.join(tmp, "tpl.xlsx")
            out = os.path.join(tmp, "out.xlsx")
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Dept: {{group.Dept}}"
            ws["B3"] = "Item: {{row.Item}}"
            ws["C5"] = "{{inc}}"
            wb.save(tpl)

            write_form_row_to_template(
                template_path=tpl,
                out_path=out,
                sheet_name="Sheet1",
                group_values={"Dept": "Sales"},
                agg_values={},
                var_values={},
                row_data={"Item": "Apple", "Dept": "Sales"},
                static_fields=[],
                inc=2,
            )

            rs = load_workbook(out)["Sheet1"]
            self.assertEqual(rs["A1"].value, "Dept: Sales")
            self.assertEqual(rs["B3"].value, "Item: Apple")
            self.assertEqual(rs["C5"].value, 2)

    def test_write_template_with_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tpl = os.path.join(tmp, "tpl.xlsx")
            out = os.path.join(tmp, "out.xlsx")
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Group: {{group.Dept}}"
            ws["A9"] = "Total: {{agg.total_qty}}"
            for col in (1, 2, 3):
                cell = ws.cell(row=10, column=col)
                cell.font = Font(bold=True, color="00FF0000")
                cell.fill = PatternFill("solid", fgColor="00DDEEFF")
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = "#,##0.00"
            ws.cell(row=10, column=1, value="{{inc}}")
            ws.cell(row=10, column=2, value="{{row.Item}}")
            ws.cell(row=10, column=3, value="{{row.Qty}}")
            wb.save(tpl)

            write_group_to_template(
                template_path=tpl,
                out_path=out,
                sheet_name="Sheet1",
                table_start_row=10,
                table_template_row=10,
                table_columns=[("Item", 2), ("Qty", 3)],
                group_df_rows=[
                    {"Item": "Apple", "Qty": 3},
                    {"Item": "Banana", "Qty": 5},
                ],
                group_values={"Dept": "Sales"},
                agg_values={"total_qty": 8},
                var_values={},
                static_fields=[],
                row_increment_col=1,
                row_increment_start=1,
            )

            self.assertTrue(os.path.isfile(out))
            result = load_workbook(out)
            rs = result["Sheet1"]
            self.assertEqual(rs["A1"].value, "Group: Sales")
            self.assertEqual(rs["A9"].value, "Total: 8")
            self.assertEqual(rs.cell(row=10, column=1).value, 1)
            self.assertEqual(rs.cell(row=10, column=2).value, "Apple")
            self.assertEqual(rs.cell(row=11, column=2).value, "Banana")
            for row in (10, 11):
                cell = rs.cell(row=row, column=2)
                self.assertTrue(cell.font.bold, f"row {row} should keep bold font")
                self.assertEqual(cell.number_format, "#,##0.00", f"row {row} number format")
                self.assertEqual(cell.fill.fgColor.rgb, "00DDEEFF", f"row {row} fill")

    def test_template_row_separate_from_start_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tpl = os.path.join(tmp, "tpl.xlsx")
            out = os.path.join(tmp, "out.xlsx")
            from openpyxl import Workbook
            from openpyxl.styles import Font

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            style_cell = ws.cell(row=8, column=2, value="{{row.Name}}")
            style_cell.font = Font(bold=True, italic=True)
            style_cell.number_format = "0.000"
            wb.save(tpl)

            write_group_to_template(
                template_path=tpl,
                out_path=out,
                sheet_name="Sheet1",
                table_start_row=10,
                table_template_row=8,
                table_columns=[("Name", 2)],
                group_df_rows=[{"Name": "One"}, {"Name": "Two"}],
                group_values={},
                agg_values={},
                var_values={},
                static_fields=[],
            )

            rs = load_workbook(out)["Sheet1"]
            self.assertEqual(rs.cell(row=8, column=2).value, "{{row.Name}}")
            for row, expected in ((10, "One"), (11, "Two")):
                cell = rs.cell(row=row, column=2)
                self.assertEqual(cell.value, expected)
                self.assertTrue(cell.font.bold)
                self.assertTrue(cell.font.italic)
                self.assertEqual(cell.number_format, "0.000")


if __name__ == "__main__":
    unittest.main()
