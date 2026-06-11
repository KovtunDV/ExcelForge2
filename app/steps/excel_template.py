from __future__ import annotations

import ast
import operator
import os
import re
import shutil
from copy import copy
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from app.steps.numeric_parse import normalize_value_for_excel, prepare_value_for_excel_cell


def safe_filename(name: str) -> str:
    name = str(name)
    name = re.sub(r'[<>:"/\\|?*]+', "_", name)
    name = name.strip().strip(".")
    return name or "output"


def parse_table_columns(raw: Any) -> list[tuple[str, int]]:
    """Список (df_col, excel_col 1-based)."""
    if raw is None or raw == "" or raw == []:
        return []
    items: list[Any]
    if isinstance(raw, dict):
        items = [{"df_col": k, "excel_col": v} for k, v in raw.items()]
    elif isinstance(raw, list):
        items = raw
    else:
        raise ValueError("table_columns must be list or dict")

    out: list[tuple[str, int]] = []
    for i, item in enumerate(items):
        if isinstance(item, dict):
            df_col = item.get("df_col") or item.get("column") or item.get("col")
            excel_col = item.get("excel_col") or item.get("excel_column")
            if df_col is None or excel_col is None:
                raise ValueError(f"table_columns[{i}]: need df_col and excel_col")
            out.append((str(df_col).strip(), int(excel_col)))
        else:
            raise ValueError(f"table_columns[{i}] must be a dict")
    return out


def parse_cell_ref(cell: str) -> tuple[int, int]:
    col_letters, row = coordinate_from_string(str(cell).strip())
    return int(row), column_index_from_string(col_letters)


_OPS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.USub: operator.neg,
    ast.Pow: operator.pow,
}


def safe_eval_scalar_expr(expr: str, names: dict[str, float]) -> float:
    """Безопасное вычисление арифметического выражения по именам agg."""
    expr = str(expr or "").strip()
    if not expr:
        raise ValueError("empty expression")

    tree = ast.parse(expr, mode="eval")

    def _eval(node: ast.AST) -> float:
        if isinstance(node, ast.Expression):
            return _eval(node.body)
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return float(node.value)
        if isinstance(node, ast.Name):
            if node.id not in names:
                raise ValueError(f"unknown name: {node.id}")
            return float(names[node.id])
        if isinstance(node, ast.BinOp):
            op_type = type(node.op)
            if op_type not in _OPS:
                raise ValueError(f"unsupported operator: {op_type.__name__}")
            return _OPS[op_type](_eval(node.left), _eval(node.right))
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, ast.USub):
            return -_eval(node.operand)
        raise ValueError(f"unsupported syntax: {type(node).__name__}")

    return _eval(tree)


def format_scalar(value: float, fmt: str | None) -> str | float | int:
    if not fmt:
        return value
    f = str(fmt).strip()
    if f.endswith("%"):
        decimals = 0
        body = f[:-1]
        if "." in body:
            decimals = len(body.split(".", 1)[1])
        return round(value * 100, decimals)
    if f.isdigit() or (f.startswith("0.") and f[1:].replace(".", "").isdigit()):
        decimals = len(f.split(".", 1)[1]) if "." in f else 0
        return round(value, decimals)
    return value


def group_key_to_str(key: Any, columns: list[str], *, separator: str = "_") -> str:
    if len(columns) == 1:
        if isinstance(key, tuple) and len(key) == 1:
            key = key[0]
        if key is None:
            return ""
        if isinstance(key, float) and key != key:
            return ""
        return str(key)
    if not isinstance(key, tuple):
        key = (key,)
    parts = []
    for v in key:
        if v is None or (isinstance(v, float) and str(v) == "nan"):
            parts.append("")
        else:
            parts.append(str(v))
    return separator.join(parts)


def build_output_filename(
    *,
    prefix: str = "",
    suffix: str = "",
    group_part: str = "",
    extension: str = ".xlsx",
    mask: str = "",
    inc: int | None = None,
    inc_position: str | None = None,
) -> str:
    ext = extension if extension.startswith(".") else f".{extension}"
    if mask:
        base = mask.format(
            group=safe_filename(group_part),
            group_key=group_part,
            inc=inc if inc is not None else "",
        )
        if not base.lower().endswith(ext.lower()):
            if not base.endswith(ext):
                base = f"{base}{ext}"
        stem, _ = os.path.splitext(safe_filename(base))
        return f"{stem}{ext}"

    if inc_position == "prefix" and inc is not None:
        inc_str = str(inc)
        prefix = f"{inc_str}_{prefix}" if prefix else f"{inc_str}_"

    body = f"{prefix}{group_part}{suffix}"
    if not body:
        body = group_part or "output"
    if inc_position == "suffix" and inc is not None:
        body = f"{body}_{inc}"
    stem = safe_filename(body)
    if stem.lower().endswith(ext.lower()):
        return stem
    return f"{stem}{ext}"


@dataclass
class _CellStyleSnapshot:
    font: Any = None
    border: Any = None
    fill: Any = None
    number_format: str = "General"
    protection: Any = None
    alignment: Any = None


@dataclass
class _RowStyleSnapshot:
    cells: dict[int, _CellStyleSnapshot] = field(default_factory=dict)
    height: float | None = None


def _capture_cell_style(cell) -> _CellStyleSnapshot:
    snap = _CellStyleSnapshot(number_format=cell.number_format or "General")
    if cell.has_style:
        snap.font = copy(cell.font)
        snap.border = copy(cell.border)
        snap.fill = copy(cell.fill)
        snap.protection = copy(cell.protection)
        snap.alignment = copy(cell.alignment)
    return snap


def _apply_cell_style(cell, snap: _CellStyleSnapshot) -> None:
    cell.number_format = snap.number_format or "General"
    if snap.font is not None:
        cell.font = copy(snap.font)
    if snap.border is not None:
        cell.border = copy(snap.border)
    if snap.fill is not None:
        cell.fill = copy(snap.fill)
    if snap.protection is not None:
        cell.protection = copy(snap.protection)
    if snap.alignment is not None:
        cell.alignment = copy(snap.alignment)


def _row_style_max_column(
    ws: Worksheet,
    row: int,
    *,
    extra_cols: set[int] | None = None,
) -> int:
    extra_cols = extra_cols or set()
    mc = max(ws.max_column or 1, max(extra_cols, default=1))
    last = 1
    for col in range(1, mc + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None or cell.has_style or col in extra_cols:
            last = max(last, col)
    return last


def _capture_row_style_snapshot(
    ws: Worksheet,
    row: int,
    *,
    extra_cols: set[int] | None = None,
) -> _RowStyleSnapshot:
    max_col = _row_style_max_column(ws, row, extra_cols=extra_cols)
    cells: dict[int, _CellStyleSnapshot] = {}
    for col in range(1, max_col + 1):
        cells[col] = _capture_cell_style(ws.cell(row=row, column=col))
    rd = ws.row_dimensions.get(row)
    height = rd.height if rd is not None else None
    return _RowStyleSnapshot(cells=cells, height=height)


def _apply_row_style_snapshot(ws: Worksheet, target_row: int, snap: _RowStyleSnapshot) -> None:
    for col, cell_snap in snap.cells.items():
        _apply_cell_style(ws.cell(row=target_row, column=col), cell_snap)
    if snap.height is not None:
        ws.row_dimensions[target_row].height = snap.height


def _apply_prepared_value(cell, value: Any, *, normalize: bool) -> None:
    prepared, fmt_override = prepare_value_for_excel_cell(
        value,
        normalize=normalize,
        number_format=cell.number_format,
    )
    cell.value = prepared
    if fmt_override:
        cell.number_format = fmt_override


def _set_cell_value(
    ws: Worksheet,
    row: int,
    col: int,
    value: Any,
    *,
    number_format: str | None = None,
    normalize: bool = False,
) -> None:
    """Записать значение, не создавая новую ячейку без стиля."""
    cell = ws.cell(row=row, column=col)
    if number_format:
        cell.number_format = number_format
    _apply_prepared_value(cell, value, normalize=normalize)


def _capture_row_placeholder_cells(ws: Worksheet, row: int) -> dict[int, str]:
    """Шаблоны {{row.*}} по столбцам строки-образца таблицы."""
    out: dict[int, str] = {}
    max_col = _row_style_max_column(ws, row)
    for col in range(1, max_col + 1):
        val = ws.cell(row=row, column=col).value
        if isinstance(val, str) and "{{row." in val:
            out[col] = val
    return out


def copy_row_styles(ws: Worksheet, src_row: int, dst_row: int, *, max_col: int | None = None) -> None:
    """Копировать оформление строки (обратная совместимость)."""
    extra: set[int] = set()
    snap = _capture_row_style_snapshot(ws, src_row, extra_cols=extra)
    if max_col is not None:
        for col in range(1, max_col + 1):
            if col not in snap.cells:
                snap.cells[col] = _capture_cell_style(ws.cell(row=src_row, column=col))
    _apply_row_style_snapshot(ws, dst_row, snap)


_PLACEHOLDER_RE = re.compile(
    r"\{\{(?:"
    r"group\.(?P<group>[^}]+)|"
    r"agg\.(?P<agg>[^}]+)|"
    r"row\.(?P<row>[^}]+)|"
    r"@(?P<var>[A-Za-z_][A-Za-z0-9_]*)|"
    r"inc(?::(?P<inc_start>\d+))?"
    r")\}\}"
)

_PURE_PLACEHOLDER_RE = re.compile(
    r"^\s*\{\{(?:"
    r"group\.(?P<group>[^}]+)|"
    r"agg\.(?P<agg>[^}]+)|"
    r"row\.(?P<row>[^}]+)|"
    r"@(?P<var>[A-Za-z_][A-Za-z0-9_]*)|"
    r"inc(?::(?P<inc_start>\d+))?"
    r")\}\}\s*$"
)


def replace_placeholders_in_text(
    text: str,
    *,
    group_values: dict[str, Any],
    agg_values: dict[str, Any],
    var_values: dict[str, Any],
    row_values: dict[str, Any] | None = None,
    inc: int | None = None,
) -> str:
    if not text or "{{" not in str(text):
        return str(text) if text is not None else ""

    def _repl(m: re.Match[str]) -> str:
        if m.group("group") is not None:
            key = m.group("group").strip()
            val = group_values.get(key, group_values.get(key.strip(), ""))
            return "" if val is None else str(val)
        if m.group("agg") is not None:
            key = m.group("agg").strip()
            val = agg_values.get(key, "")
            return "" if val is None else str(val)
        if m.group("var") is not None:
            key = m.group("var").strip()
            val = var_values.get(key, var_values.get(f"@{key}", ""))
            return "" if val is None else str(val)
        if m.group("row") is not None:
            if row_values is None:
                return m.group(0)
            key = m.group("row").strip()
            val = row_values.get(key, "")
            return "" if val is None else str(val)
        if m.group(0).startswith("{{inc"):
            start = int(m.group("inc_start") or "1")
            if inc is None:
                return m.group(0)
            return str(start + inc - 1)

        return m.group(0)

    return _PLACEHOLDER_RE.sub(_repl, str(text))


def resolve_cell_value(
    text: str,
    *,
    group_values: dict[str, Any],
    agg_values: dict[str, Any],
    var_values: dict[str, Any],
    row_values: dict[str, Any] | None = None,
    inc: int | None = None,
    number_format: str | None = None,
    normalize: bool = False,
) -> Any:
    """Подстановка в ячейку; при normalize=True чистый agg/row может стать числом."""
    raw = str(text) if text is not None else ""
    m = _PURE_PLACEHOLDER_RE.match(raw)
    if m:
        if m.group("agg") is not None:
            key = m.group("agg").strip()
            val = agg_values.get(key, "")
            if normalize:
                return normalize_value_for_excel(val, number_format=number_format)
            return val
        if m.group("row") is not None and row_values is not None:
            key = m.group("row").strip()
            val = row_values.get(key, "")
            if normalize:
                return normalize_value_for_excel(val, number_format=number_format)
            return val
        if m.group("group") is not None:
            key = m.group("group").strip()
            val = group_values.get(key, group_values.get(key.strip(), ""))
            return "" if val is None else val
        if m.group("var") is not None:
            key = m.group("var").strip()
            val = var_values.get(key, var_values.get(f"@{key}", ""))
            if normalize:
                num = normalize_value_for_excel(val)
                return num if isinstance(num, (int, float)) and not isinstance(num, bool) else val
            return val
        if m.group(0).lstrip().startswith("{{inc"):
            start = int(m.group("inc_start") or "1")
            if inc is not None:
                return start + inc - 1
    return replace_placeholders_in_text(
        raw,
        group_values=group_values,
        agg_values=agg_values,
        var_values=var_values,
        row_values=row_values,
        inc=inc,
    )


def apply_static_fields(
    ws: Worksheet,
    fields: list[dict[str, Any]],
    context: dict[str, Any],
    *,
    row_values: dict[str, Any] | None = None,
    inc: int | None = None,
    normalize: bool = False,
) -> None:
    for i, item in enumerate(fields):
        if not isinstance(item, dict):
            raise ValueError(f"static_fields[{i}] must be a dict")
        cell_ref = item.get("cell")
        if not cell_ref:
            raise ValueError(f"static_fields[{i}]: cell is required")
        row, col = parse_cell_ref(str(cell_ref))
        raw_value = item.get("value", "")
        cell = ws.cell(row=row, column=col)
        resolved = resolve_cell_value(
            str(raw_value),
            group_values=context.get("group", {}),
            agg_values=context.get("agg", {}),
            var_values=context.get("var", {}),
            row_values=row_values,
            inc=inc,
            number_format=cell.number_format,
            normalize=normalize,
        )
        _apply_prepared_value(cell, resolved, normalize=normalize)


def replace_in_worksheet(
    ws: Worksheet,
    *,
    group_values: dict[str, Any],
    agg_values: dict[str, Any],
    var_values: dict[str, Any],
    skip_rows: set[int] | None = None,
    only_rows: set[int] | None = None,
    row_data: dict[str, Any] | None = None,
    inc: int | None = None,
    normalize: bool = False,
) -> None:
    skip_rows = skip_rows or set()
    for row in ws.iter_rows():
        r_idx = row[0].row
        if r_idx in skip_rows:
            continue
        if only_rows is not None and r_idx not in only_rows:
            continue
        for cell in row:
            val = cell.value
            if not isinstance(val, str) or "{{" not in val:
                continue
            resolved = resolve_cell_value(
                val,
                group_values=group_values,
                agg_values=agg_values,
                var_values=var_values,
                row_values=row_data,
                inc=inc,
                number_format=cell.number_format,
                normalize=normalize,
            )
            _apply_prepared_value(cell, resolved, normalize=normalize)


def _resolve_worksheet(wb, sheet_name: str) -> Worksheet:
    if sheet_name and sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.active


def write_group_to_template(
    *,
    template_path: str,
    out_path: str,
    sheet_name: str,
    table_start_row: int,
    table_template_row: int | None,
    table_columns: list[tuple[str, int]],
    group_df_rows: list[dict[str, Any]],
    group_values: dict[str, Any],
    agg_values: dict[str, Any],
    var_values: dict[str, Any],
    static_fields: list[dict[str, Any]],
    row_increment_col: int | None = None,
    row_increment_start: int = 1,
    normalize_values: bool = False,
) -> None:
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    shutil.copy2(template_path, out_path)
    wb = load_workbook(out_path)
    ws = _resolve_worksheet(wb, sheet_name)

    start_row = max(1, int(table_start_row))
    style_row = max(1, int(table_template_row or table_start_row))

    style_cols: set[int] = {col for _, col in table_columns}
    if row_increment_col is not None:
        style_cols.add(int(row_increment_col))
    style_snapshot = _capture_row_style_snapshot(ws, style_row, extra_cols=style_cols)
    row_placeholder_templates = _capture_row_placeholder_cells(ws, style_row)

    ctx = {"group": group_values, "agg": agg_values, "var": var_values}
    if static_fields:
        apply_static_fields(ws, static_fields, ctx, normalize=normalize_values)

    n = len(group_df_rows)
    output_rows = {start_row + i for i in range(max(n, 1))}

    replace_in_worksheet(
        ws,
        group_values=group_values,
        agg_values=agg_values,
        var_values=var_values,
        skip_rows=output_rows,
        normalize=normalize_values,
    )

    if n > 1:
        ws.insert_rows(start_row + 1, amount=n - 1)

    for i, row_dict in enumerate(group_df_rows):
        excel_row = start_row + i
        inc_num = i + 1

        if excel_row != style_row or i > 0:
            _apply_row_style_snapshot(ws, excel_row, style_snapshot)

        if row_increment_col is not None:
            inc_snap = style_snapshot.cells.get(int(row_increment_col))
            _set_cell_value(
                ws,
                excel_row,
                row_increment_col,
                row_increment_start + i,
                number_format=inc_snap.number_format if inc_snap else None,
                normalize=normalize_values,
            )

        if i > 0:
            for col, template_text in row_placeholder_templates.items():
                ws.cell(row=excel_row, column=col, value=template_text)

        for df_col, excel_col in table_columns:
            if df_col not in row_dict:
                continue
            val = row_dict[df_col]
            if val is None or (isinstance(val, float) and val != val):
                cell_val = None
            else:
                cell_val = val
            col_snap = style_snapshot.cells.get(excel_col)
            _set_cell_value(
                ws,
                excel_row,
                excel_col,
                cell_val,
                number_format=col_snap.number_format if col_snap else None,
                normalize=normalize_values,
            )

        replace_in_worksheet(
            ws,
            group_values=group_values,
            agg_values=agg_values,
            var_values=var_values,
            only_rows={excel_row},
            row_data=row_dict,
            inc=inc_num,
            normalize=normalize_values,
        )

    footer_start = start_row + max(n, 1)
    replace_in_worksheet(
        ws,
        group_values=group_values,
        agg_values=agg_values,
        var_values=var_values,
        only_rows={r for r in range(footer_start, (ws.max_row or footer_start) + 1)},
        normalize=normalize_values,
    )

    wb.save(out_path)


def write_form_row_to_template(
    *,
    template_path: str,
    out_path: str,
    sheet_name: str,
    group_values: dict[str, Any],
    agg_values: dict[str, Any],
    var_values: dict[str, Any],
    row_data: dict[str, Any],
    static_fields: list[dict[str, Any]],
    inc: int | None = None,
    normalize_values: bool = False,
) -> None:
    """
    Однострочный режим: {{row.*}}, {{group.*}}, {{agg.*}}, {{@var}}, {{inc}}
    подставляются по всему листу; отдельный файл на каждую строку DF.
    """
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    shutil.copy2(template_path, out_path)
    wb = load_workbook(out_path)
    ws = _resolve_worksheet(wb, sheet_name)

    ctx = {"group": group_values, "agg": agg_values, "var": var_values}
    if static_fields:
        apply_static_fields(
            ws, static_fields, ctx, row_values=row_data, inc=inc, normalize=normalize_values
        )

    replace_in_worksheet(
        ws,
        group_values=group_values,
        agg_values=agg_values,
        var_values=var_values,
        row_data=row_data,
        inc=inc,
        normalize=normalize_values,
    )

    wb.save(out_path)
